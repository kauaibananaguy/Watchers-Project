#!/usr/bin/env python3
"""Build a complete GEIPAN source-acquisition snapshot from official exports.

The package preserves both the current live XLSX exports and the older, much
wider CSV exports. Overlapping rows remain separate source records and are
connected through reconciliation tables. No source row is discarded and no
canonical Atlas merge is performed here.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import shutil
import sqlite3
import time
import urllib.error
import urllib.request
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

COLLECTION_ID = "SRC-COLLECTION-GEIPAN"
PACKAGE_ID = "UFO-ATLAS-IMPORT-GEIPAN-SOURCE-SNAPSHOT-0.2.0"
URLS = {
    "GEIPAN_current_cases.xlsx": "https://www.geipan.fr/fr/cnes/export/cas",
    "GEIPAN_current_testimonies.xlsx": "https://www.geipan.fr/fr/cnes/export/temoignages",
    "GEIPAN_legacy_cases_35_fields.csv": "https://www.cnes-geipan.fr/sites/default/files/Base_de_donn%C3%A9es_des_cas.csv",
    "GEIPAN_legacy_testimonies_263_fields.csv": "https://www.cnes-geipan.fr/sites/default/files/Base_de_donn%C3%A9es_des_t%C3%A9moignages.csv",
    "GEIPAN_database_update_page.html": "https://www.geipan.fr/fr/actualites/mise-a-jour-csv",
    "GEIPAN_statistics_page.html": "https://www.geipan.fr/fr/stats",
    "GEIPAN_field_description_2019-01-07.xlsx": "https://www.cnes-geipan.fr/sites/default/files/Description_des_tables_et_champs_de_donn%C3%A9es_de_la_base_du_geipan_2019-01-07.xlsx",
    "GEIPAN_database_history_2019-02-26.pdf": "https://www.cnes-geipan.fr/sites/default/files/2019-02-26_Historique_des_bases_au_GEIPAN.pdf",
}


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value).strip()
    return value


def json_text(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)


def fetch(url: str, referer: str | None = None, retries: int = 12) -> tuple[bytes, str, dict[str, str]]:
    last: Exception | None = None
    for attempt in range(retries):
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (compatible; Watchers-UFO-Atlas/1.0; public-source-preservation)",
                "Accept": "*/*",
                "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.5",
                "Connection": "close",
            }
            if referer:
                headers["Referer"] = referer
            request = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(request, timeout=300) as response:
                data = response.read()
                if not data:
                    raise RuntimeError("empty response")
                return data, response.geturl(), dict(response.headers.items())
        except Exception as exc:  # noqa: BLE001
            last = exc
            if attempt + 1 < retries:
                if isinstance(exc, urllib.error.HTTPError) and exc.code == 429:
                    retry_after = exc.headers.get("Retry-After") if exc.headers else None
                    wait = max(90, int(retry_after)) if retry_after and retry_after.isdigit() else min(600, 120 + attempt * 30)
                else:
                    wait = min(60, 2**attempt)
                print(f"Retrying {url} in {wait}s after {type(exc).__name__}: {exc}", flush=True)
                time.sleep(wait)
    raise RuntimeError(f"Unable to fetch {url}: {last}")


def decode_csv(path: Path) -> tuple[str, str, csv.Dialect]:
    data = path.read_bytes()
    text: str | None = None
    encoding = ""
    for candidate in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = data.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = data.decode("utf-8", errors="replace")
        encoding = "utf-8-replace"
    try:
        dialect = csv.Sniffer().sniff(text[:200000], delimiters=";,\t|")
    except csv.Error:
        class Semicolon(csv.excel):
            delimiter = ";"
        dialect = Semicolon()
    return text, encoding, dialect


def read_csv_rows(path: Path) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    text, encoding, dialect = decode_csv(path)
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = [str(value) for value in (reader.fieldnames or [])]
    rows: list[dict[str, Any]] = []
    blank_rows = 0
    for source_row_number, source_row in enumerate(reader, start=2):
        row = {str(key): normalize_scalar(value) for key, value in source_row.items() if key is not None}
        if not any(value not in (None, "") for value in row.values()):
            blank_rows += 1
            continue
        rows.append({"source_row_number": source_row_number, "values": row})
    return headers, rows, {
        "encoding": encoding,
        "delimiter": dialect.delimiter,
        "quotechar": dialect.quotechar,
        "header_count": len(headers),
        "row_count": len(rows),
        "blank_rows_skipped": blank_rows,
    }


def read_xlsx_rows(path: Path) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    header_row = next(iterator)
    headers = [str(value).strip() if value is not None else f"UNNAMED_COLUMN_{index + 1}" for index, value in enumerate(header_row)]
    rows: list[dict[str, Any]] = []
    blank_rows = 0
    for source_row_number, values in enumerate(iterator, start=2):
        padded = list(values) + [None] * max(0, len(headers) - len(values))
        row = {headers[index]: normalize_scalar(padded[index]) for index in range(len(headers))}
        if not any(value not in (None, "") for value in row.values()):
            blank_rows += 1
            continue
        rows.append({"source_row_number": source_row_number, "values": row})
    return headers, rows, {
        "sheet_name": worksheet.title,
        "header_count": len(headers),
        "row_count": len(rows),
        "blank_rows_skipped": blank_rows,
    }


def parse_live_stats(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8", errors="replace")
    total_match = re.search(
        r'class="stat-total"[^>]*>\s*([0-9][0-9\s]*)\s*Cas\s*-\s*Statistiques\s+du\s+([0-9/]+)',
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    classes: dict[str, int] = {}
    for count, code in re.findall(
        r'class="stat-legende-number"[^>]*>\s*\(([0-9][0-9\s]*)\s*Cas\s*([ABCD])\)',
        text,
        flags=re.IGNORECASE | re.DOTALL,
    ):
        classes[code.upper()] = int(re.sub(r"\s+", "", count))
    return {
        "published_case_count": int(re.sub(r"\s+", "", total_match.group(1))) if total_match else None,
        "statistics_date_original": total_match.group(2) if total_match else None,
        "classification_counts": classes,
    }


def stable_title_key(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text).strip().casefold()
    return text


def write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True, default=str) + "\n")


def build(args: argparse.Namespace) -> None:
    root = Path(args.output_root)
    if root.exists():
        shutil.rmtree(root)
    raw = root / "raw"
    output = root / "output"
    release = root / "release"
    raw.mkdir(parents=True)
    output.mkdir()
    release.mkdir()
    acquired_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

    file_inventory: list[dict[str, Any]] = []
    for index, (file_name, url) in enumerate(URLS.items(), start=1):
        referer = "https://www.geipan.fr/fr/recherche/cas" if "/fr/cnes/export/" in url else None
        print(f"Acquiring {file_name}", flush=True)
        data, final_url, headers = fetch(url, referer=referer)
        path = raw / file_name
        path.write_bytes(data)
        file_inventory.append(
            {
                "file_id": f"GEIPAN-FILE-{index:03d}",
                "file_name": file_name,
                "requested_url": url,
                "final_url": final_url,
                "bytes": len(data),
                "sha256": sha256_bytes(data),
                "http_headers": headers,
                "acquired_at": acquired_at,
            }
        )
    file_by_name = {row["file_name"]: row for row in file_inventory}

    table_specs = [
        ("GEIPAN_current_cases.xlsx", "CURRENT_CASE_EXPORT", read_xlsx_rows),
        ("GEIPAN_current_testimonies.xlsx", "CURRENT_TESTIMONY_EXPORT", read_xlsx_rows),
        ("GEIPAN_legacy_cases_35_fields.csv", "LEGACY_CASE_EXPORT", read_csv_rows),
        ("GEIPAN_legacy_testimonies_263_fields.csv", "LEGACY_TESTIMONY_EXPORT", read_csv_rows),
    ]
    parsed_tables: dict[str, dict[str, Any]] = {}
    source_fields: list[dict[str, Any]] = []
    ledger: list[dict[str, Any]] = []
    sequence = 0
    seen_source_ids: Counter[str] = Counter()

    for file_name, family, reader_function in table_specs:
        headers, rows, profile = reader_function(raw / file_name)
        parsed_tables[family] = {"file_name": file_name, "headers": headers, "rows": rows, "profile": profile}
        for ordinal, field_name in enumerate(headers, start=1):
            source_fields.append(
                {
                    "source_file_id": file_by_name[file_name]["file_id"],
                    "source_record_family": family,
                    "field_ordinal": ordinal,
                    "source_field_name": field_name,
                    "mapping_status": "PENDING_GMR_MAPPING",
                }
            )
        for row in rows:
            sequence += 1
            values = row["values"]
            if family in {"CURRENT_CASE_EXPORT", "LEGACY_CASE_EXPORT"}:
                native_id = str(values.get("ID Etude de Cas") or values.get("cas_numEtude") or "").strip()
                title = str(values.get("Titre du Cas") or values.get("cas_nom_dossier") or "").strip()
                record_kind = "CASE"
            elif family == "CURRENT_TESTIMONY_EXPORT":
                native_id = ""
                title = str(values.get("Témoignage") or "").strip()
                record_kind = "TESTIMONY_OBSERVATION"
            else:
                native_id = str(values.get("id_temoignage") or "").strip()
                title = str(values.get("tem_nom_dossier") or "").strip()
                record_kind = "TESTIMONY_OBSERVATION"

            if native_id:
                base_source_id = f"GEIPAN-{family}-{native_id}"
            else:
                title_hash = hashlib.sha256(stable_title_key(title).encode("utf-8")).hexdigest()[:20]
                base_source_id = f"GEIPAN-{family}-TITLE-{title_hash}"
            seen_source_ids[base_source_id] += 1
            source_record_id = base_source_id if seen_source_ids[base_source_id] == 1 else f"{base_source_id}-DUP-{seen_source_ids[base_source_id]:03d}"
            raw_payload = json_text(values)
            ledger.append(
                {
                    "source_collection_id": COLLECTION_ID,
                    "source_sequence": sequence,
                    "source_record_id": source_record_id,
                    "source_record_family": family,
                    "record_kind": record_kind,
                    "source_file_id": file_by_name[file_name]["file_id"],
                    "source_row_number": row["source_row_number"],
                    "source_locator": f"{file_name}:row:{row['source_row_number']}",
                    "source_native_id": native_id or None,
                    "source_title": title or None,
                    "raw_payload": raw_payload,
                    "row_sha256": sha256_bytes(raw_payload.encode("utf-8")),
                    "transformation_status": "ACQUIRED_UNTRANSFORMED",
                }
            )

    ledger_by_family: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in ledger:
        ledger_by_family[row["source_record_family"]].append(row)

    current_case = {row["source_native_id"]: row for row in ledger_by_family["CURRENT_CASE_EXPORT"] if row["source_native_id"]}
    legacy_case = {row["source_native_id"]: row for row in ledger_by_family["LEGACY_CASE_EXPORT"] if row["source_native_id"]}
    case_reconciliation: list[dict[str, Any]] = []
    for case_id in sorted(set(current_case) | set(legacy_case)):
        current = current_case.get(case_id)
        legacy = legacy_case.get(case_id)
        case_reconciliation.append(
            {
                "geipan_case_id": case_id,
                "current_source_record_id": current["source_record_id"] if current else None,
                "legacy_source_record_id": legacy["source_record_id"] if legacy else None,
                "reconciliation_status": "CURRENT_AND_LEGACY" if current and legacy else "CURRENT_ONLY" if current else "LEGACY_ONLY",
                "proposed_resolution": "PROPOSE_SOURCE_VARIANT" if current and legacy else "PROPOSE_NEW_CANONICAL_OR_MATCH_MASTER",
            }
        )

    current_test_by_title: dict[str, list[dict[str, Any]]] = defaultdict(list)
    legacy_test_by_title: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in ledger_by_family["CURRENT_TESTIMONY_EXPORT"]:
        current_test_by_title[stable_title_key(row["source_title"])].append(row)
    for row in ledger_by_family["LEGACY_TESTIMONY_EXPORT"]:
        legacy_test_by_title[stable_title_key(row["source_title"])].append(row)
    testimony_reconciliation: list[dict[str, Any]] = []
    for title_key in sorted(set(current_test_by_title) | set(legacy_test_by_title)):
        current_rows = current_test_by_title.get(title_key, [])
        legacy_rows = legacy_test_by_title.get(title_key, [])
        title = (current_rows or legacy_rows)[0]["source_title"]
        testimony_reconciliation.append(
            {
                "normalized_title_key": title_key,
                "representative_title": title,
                "current_source_record_ids": [row["source_record_id"] for row in current_rows],
                "legacy_source_record_ids": [row["source_record_id"] for row in legacy_rows],
                "current_count": len(current_rows),
                "legacy_count": len(legacy_rows),
                "reconciliation_status": "CURRENT_AND_LEGACY" if current_rows and legacy_rows else "CURRENT_ONLY" if current_rows else "LEGACY_ONLY",
                "proposed_resolution": "PROPOSE_SOURCE_VARIANT" if current_rows and legacy_rows else "PROPOSE_NEW_CANONICAL_OR_MATCH_MASTER",
            }
        )

    live_stats = parse_live_stats(raw / "GEIPAN_statistics_page.html")
    family_counts = Counter(row["source_record_family"] for row in ledger)
    case_status_counts = Counter(row["reconciliation_status"] for row in case_reconciliation)
    testimony_status_counts = Counter(row["reconciliation_status"] for row in testimony_reconciliation)

    source_collection = {
        "source_collection_id": COLLECTION_ID,
        "collection_title": "GEIPAN published case and testimony database exports",
        "attribution_name": "GEIPAN — Groupe d’études et d’informations sur les phénomènes aérospatiaux non identifiés",
        "owner_custodian": "Centre national d’études spatiales (CNES)",
        "source_type": "OFFICIAL_GOVERNMENT_DATABASE_EXPORT",
        "acquisition_method": "Direct public download from official GEIPAN/CNES endpoints",
        "acquired_at": acquired_at,
        "gmr_version": "UFO Atlas GMR v1.0.0",
        "integration_standard": "UFO-ATLAS-INT-STD-1.0.0",
        "preliminary_master_reference": "UFO Atlas v0.6.0 at issuance; rebase against latest verified master at handoff",
        "public_silo_flag": False,
        "rights_and_privacy_notes": (
            "GEIPAN publishes anonymized records and exports for public use. Preserve attribution, original French wording, "
            "source identifiers, and the source's coordinate/age generalization. Do not attempt to reverse anonymization."
        ),
        "source_file_count": len(file_inventory),
        "source_record_count": len(ledger),
        "current_case_export_rows": family_counts["CURRENT_CASE_EXPORT"],
        "current_testimony_export_rows": family_counts["CURRENT_TESTIMONY_EXPORT"],
        "legacy_case_export_rows": family_counts["LEGACY_CASE_EXPORT"],
        "legacy_testimony_export_rows": family_counts["LEGACY_TESTIMONY_EXPORT"],
        "unique_case_ids_across_exports": len(case_reconciliation),
        "unique_testimony_title_keys_across_exports": len(testimony_reconciliation),
        "live_stats": live_stats,
    }

    # Machine-readable outputs.
    (output / "SOURCE_COLLECTION.json").write_text(json.dumps(source_collection, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8")
    (output / "SOURCE_FILE_INVENTORY.json").write_text(json.dumps(file_inventory, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8")
    (output / "SOURCE_TABLE_PROFILES.json").write_text(
        json.dumps({family: {"file_name": value["file_name"], "headers": value["headers"], "profile": value["profile"]} for family, value in parsed_tables.items()}, indent=2, ensure_ascii=False, default=str) + "\n",
        encoding="utf-8",
    )
    write_jsonl(output / "SOURCE_RECORD_LEDGER.jsonl", ledger)
    write_jsonl(output / "CASE_EXPORT_RECONCILIATION.jsonl", case_reconciliation)
    write_jsonl(output / "TESTIMONY_EXPORT_RECONCILIATION.jsonl", testimony_reconciliation)
    with (output / "SOURCE_RECORD_LEDGER.csv").open("w", encoding="utf-8", newline="") as handle:
        columns = list(ledger[0])
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(ledger)
    with (output / "SOURCE_FIELD_INVENTORY.csv").open("w", encoding="utf-8", newline="") as handle:
        columns = list(source_fields[0])
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(source_fields)

    database = output / "GEIPAN_SOURCE_SNAPSHOT_v0.2.0.sqlite"
    connection = sqlite3.connect(database)
    connection.execute("PRAGMA foreign_keys=ON")
    connection.executescript(
        """
        CREATE TABLE source_collection(source_collection_id TEXT PRIMARY KEY, payload_json TEXT NOT NULL);
        CREATE TABLE source_file(
          file_id TEXT PRIMARY KEY, file_name TEXT NOT NULL UNIQUE, requested_url TEXT NOT NULL,
          final_url TEXT NOT NULL, bytes INTEGER NOT NULL, sha256 TEXT NOT NULL,
          http_headers_json TEXT NOT NULL, acquired_at TEXT NOT NULL
        );
        CREATE TABLE source_record_ledger(
          source_collection_id TEXT NOT NULL REFERENCES source_collection(source_collection_id),
          source_sequence INTEGER NOT NULL UNIQUE, source_record_id TEXT PRIMARY KEY,
          source_record_family TEXT NOT NULL, record_kind TEXT NOT NULL,
          source_file_id TEXT NOT NULL REFERENCES source_file(file_id), source_row_number INTEGER NOT NULL,
          source_locator TEXT NOT NULL, source_native_id TEXT, source_title TEXT,
          raw_payload TEXT NOT NULL, row_sha256 TEXT NOT NULL, transformation_status TEXT NOT NULL
        );
        CREATE TABLE source_field_inventory(
          source_file_id TEXT NOT NULL REFERENCES source_file(file_id), source_record_family TEXT NOT NULL,
          field_ordinal INTEGER NOT NULL, source_field_name TEXT NOT NULL, mapping_status TEXT NOT NULL,
          PRIMARY KEY(source_file_id, field_ordinal)
        );
        CREATE TABLE case_export_reconciliation(
          geipan_case_id TEXT PRIMARY KEY, current_source_record_id TEXT REFERENCES source_record_ledger(source_record_id),
          legacy_source_record_id TEXT REFERENCES source_record_ledger(source_record_id),
          reconciliation_status TEXT NOT NULL, proposed_resolution TEXT NOT NULL
        );
        CREATE TABLE testimony_export_reconciliation(
          normalized_title_key TEXT PRIMARY KEY, representative_title TEXT,
          current_source_record_ids_json TEXT NOT NULL, legacy_source_record_ids_json TEXT NOT NULL,
          current_count INTEGER NOT NULL, legacy_count INTEGER NOT NULL,
          reconciliation_status TEXT NOT NULL, proposed_resolution TEXT NOT NULL
        );
        CREATE INDEX idx_geipan_ledger_family ON source_record_ledger(source_record_family);
        CREATE INDEX idx_geipan_ledger_native ON source_record_ledger(source_native_id);
        CREATE INDEX idx_geipan_ledger_title ON source_record_ledger(source_title);
        CREATE INDEX idx_geipan_case_status ON case_export_reconciliation(reconciliation_status);
        CREATE INDEX idx_geipan_test_status ON testimony_export_reconciliation(reconciliation_status);
        """
    )
    connection.execute("INSERT INTO source_collection VALUES (?,?)", (COLLECTION_ID, json_text(source_collection)))
    connection.executemany(
        "INSERT INTO source_file VALUES (?,?,?,?,?,?,?,?)",
        [
            (
                row["file_id"], row["file_name"], row["requested_url"], row["final_url"], row["bytes"],
                row["sha256"], json_text(row["http_headers"]), row["acquired_at"],
            )
            for row in file_inventory
        ],
    )
    connection.executemany(
        "INSERT INTO source_record_ledger VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        [
            (
                row["source_collection_id"], row["source_sequence"], row["source_record_id"],
                row["source_record_family"], row["record_kind"], row["source_file_id"],
                row["source_row_number"], row["source_locator"], row["source_native_id"],
                row["source_title"], row["raw_payload"], row["row_sha256"], row["transformation_status"],
            )
            for row in ledger
        ],
    )
    connection.executemany(
        "INSERT INTO source_field_inventory VALUES (?,?,?,?,?)",
        [(row["source_file_id"], row["source_record_family"], row["field_ordinal"], row["source_field_name"], row["mapping_status"]) for row in source_fields],
    )
    connection.executemany(
        "INSERT INTO case_export_reconciliation VALUES (?,?,?,?,?)",
        [(row["geipan_case_id"], row["current_source_record_id"], row["legacy_source_record_id"], row["reconciliation_status"], row["proposed_resolution"]) for row in case_reconciliation],
    )
    connection.executemany(
        "INSERT INTO testimony_export_reconciliation VALUES (?,?,?,?,?,?,?,?)",
        [
            (
                row["normalized_title_key"], row["representative_title"], json_text(row["current_source_record_ids"]),
                json_text(row["legacy_source_record_ids"]), row["current_count"], row["legacy_count"],
                row["reconciliation_status"], row["proposed_resolution"],
            )
            for row in testimony_reconciliation
        ],
    )
    connection.commit()
    quick = connection.execute("PRAGMA quick_check").fetchone()[0]
    foreign_keys = len(connection.execute("PRAGMA foreign_key_check").fetchall())
    ledger_count = connection.execute("SELECT COUNT(*) FROM source_record_ledger").fetchone()[0]
    min_seq, max_seq, distinct_seq = connection.execute("SELECT MIN(source_sequence), MAX(source_sequence), COUNT(DISTINCT source_sequence) FROM source_record_ledger").fetchone()
    duplicate_ids = connection.execute("SELECT COUNT(*) FROM (SELECT source_record_id FROM source_record_ledger GROUP BY source_record_id HAVING COUNT(*)>1)").fetchone()[0]
    connection.close()

    validation = {
        "overall_status": "PASS" if (
            quick == "ok" and foreign_keys == 0 and ledger_count == len(ledger)
            and min_seq == 1 and max_seq == len(ledger) and distinct_seq == len(ledger)
            and duplicate_ids == 0 and family_counts["CURRENT_CASE_EXPORT"] == live_stats.get("published_case_count")
        ) else "FAIL",
        "package_id": PACKAGE_ID,
        "stage": "CURRENT_AND_LEGACY_OFFICIAL_SOURCE_ACQUISITION",
        "source_collection_id": COLLECTION_ID,
        "acquired_at": acquired_at,
        "gmr_version": "UFO Atlas GMR v1.0.0",
        "source_record_count": len(ledger),
        "source_record_counts_by_family": dict(family_counts),
        "source_field_instances": len(source_fields),
        "current_case_export_rows": family_counts["CURRENT_CASE_EXPORT"],
        "live_stats_published_cases": live_stats.get("published_case_count"),
        "current_case_export_matches_live_stats": family_counts["CURRENT_CASE_EXPORT"] == live_stats.get("published_case_count"),
        "unique_case_ids_across_exports": len(case_reconciliation),
        "case_reconciliation_status_counts": dict(case_status_counts),
        "unique_testimony_title_keys_across_exports": len(testimony_reconciliation),
        "testimony_reconciliation_status_counts": dict(testimony_status_counts),
        "sqlite_quick_check": quick,
        "foreign_key_violations": foreign_keys,
        "duplicate_source_record_ids": duplicate_ids,
        "source_sequence_start": min_seq,
        "source_sequence_end": max_seq,
        "source_sequence_distinct_count": distinct_seq,
        "candidate_records_built": 0,
        "gmr_mapping_status": "NEXT_STAGE",
        "next_stage": "SOURCE_FIELD_MAPPING_GMR_TRANSFORMATION_CASE_PACKET_AND_ATTACHMENT_ACQUISITION",
        "limitations": [
            "This package is an acquisition and reconciliation checkpoint, not the final GMR import module.",
            "Current XLSX exports are complete for the published case/testimony tables but contain fewer columns than the older detailed CSV exports; both are preserved.",
            "The official 2019 field-description workbook is retained even if its malformed ZIP offsets prevent ordinary workbook parsing.",
            "Case pages, PDFs, testimony documents, photographs, diagrams, and other attachments remain to be acquired and bound before final handoff.",
        ],
    }
    (output / "VALIDATION_REPORT.json").write_text(json.dumps(validation, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    if validation["overall_status"] != "PASS":
        raise RuntimeError(json.dumps(validation, indent=2, ensure_ascii=False))

    readme = f"""# GEIPAN official source snapshot — v0.2.0

Status: **PASS**  
Source collection: `{COLLECTION_ID}`  
Acquired: `{acquired_at}`

This checkpoint combines GEIPAN's current live XLSX exports with its older, wider CSV exports. Every physical row is preserved as a separate source record. Overlapping current and legacy rows are reconciled but never collapsed or discarded.

## Source rows preserved

- Current case export: **{family_counts['CURRENT_CASE_EXPORT']:,}**
- Current testimony export: **{family_counts['CURRENT_TESTIMONY_EXPORT']:,}**
- Legacy detailed case export: **{family_counts['LEGACY_CASE_EXPORT']:,}**
- Legacy detailed testimony/observation export: **{family_counts['LEGACY_TESTIMONY_EXPORT']:,}**
- Total uninterrupted source ledger: **{len(ledger):,}**

## Reconciliation

- Unique case IDs across both case exports: **{len(case_reconciliation):,}**
- Present in both: **{case_status_counts['CURRENT_AND_LEGACY']:,}**
- Current only: **{case_status_counts['CURRENT_ONLY']:,}**
- Legacy only: **{case_status_counts['LEGACY_ONLY']:,}**
- Unique normalized testimony-title keys: **{len(testimony_reconciliation):,}**

The current case XLSX count matches the current GEIPAN published-case count of **{live_stats.get('published_case_count'):,}** reported on the official statistics page dated **{live_stats.get('statistics_date_original')}**.

## Architectural boundary

This is a source-preservation module for the one source-neutral UFO Atlas, not a public GEIPAN database. Canonical candidates, complete chronological narratives, GMR values, typed relationships, and final master-match proposals are built in the next stage.
"""
    (output / "README_FIRST.md").write_text(readme, encoding="utf-8")

    # Package raw official files and output evidence without nested ZIP files.
    shutil.copytree(raw, release / "raw")
    shutil.copytree(output, release / "output")
    manifest_files = []
    for path in sorted(release.rglob("*")):
        if path.is_file() and path.name != "manifest.json":
            manifest_files.append({"path": str(path.relative_to(release)), "bytes": path.stat().st_size, "sha256": sha256_file(path)})
    manifest = {
        "package_id": PACKAGE_ID,
        "source_collection_id": COLLECTION_ID,
        "created_at": acquired_at,
        "files": manifest_files,
        "manifest_note": "Manifest excludes its own self-hash.",
    }
    manifest_text = json.dumps(manifest, indent=2, ensure_ascii=False) + "\n"
    (release / "output" / "manifest.json").write_text(manifest_text, encoding="utf-8")
    (output / "manifest.json").write_text(manifest_text, encoding="utf-8")

    package = Path(args.package)
    package.unlink(missing_ok=True)
    with zipfile.ZipFile(package, "w", zipfile.ZIP_DEFLATED, allowZip64=True) as archive:
        for path in sorted(release.rglob("*")):
            if path.is_file():
                archive.write(path, path.relative_to(release))
    with zipfile.ZipFile(package) as archive:
        bad = archive.testzip()
        nested = [name for name in archive.namelist() if name.lower().endswith(".zip")]
    if bad or nested:
        raise RuntimeError(f"Package validation failed: bad={bad}, nested={nested}")
    Path(str(package) + ".sha256").write_text(f"{sha256_file(package)}  {package.name}\n", encoding="utf-8")
    print(json.dumps(validation, indent=2, ensure_ascii=False))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-root", default="geipan_source_snapshot")
    parser.add_argument("--package", default="UFO_ATLAS_IMPORT_GEIPAN_GMR1.0.0_v0.2.0_SOURCE_SNAPSHOT.zip")
    build(parser.parse_args())


if __name__ == "__main__":
    main()
