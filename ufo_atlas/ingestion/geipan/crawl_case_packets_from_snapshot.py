#!/usr/bin/env python3
"""Plan GEIPAN case-packet acquisition directly from the validated source snapshot.

The current official GEIPAN case workbook does not embed hyperlinks.  Its
``ID Etude de Cas`` column is, however, the stable path component used by the
public case pages.  This planner therefore preserves any explicit URLs it can
find and deterministically derives the missing case-page URLs from that column.
"""
from __future__ import annotations

import argparse
import json
import re
import unicodedata
import urllib.parse
from pathlib import Path

from openpyxl import load_workbook

import crawl_case_packets as base


CASE_ID_RE = re.compile(r"^\d{4}-\d{2}-\d{5}$")
CASE_URL_ROOT = "https://www.geipan.fr/fr/cas/"


def normalized_header(value: object) -> str:
    text = base.clean(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def discover_with_workbooks(root: Path) -> list[dict[str, str]]:
    found = {row["case_url"]: row for row in base.discover_urls(root)}

    def add(url: str, source: str, source_id: str = "") -> None:
        url = base.normalize_url(url)
        if not re.search(r"/(?:fr|en)/(?:cas|case)/", url, flags=re.I):
            return
        found.setdefault(
            url,
            {
                "case_url": url,
                "source_locator": source,
                "source_case_id": source_id,
            },
        )

    url_rx = re.compile(r"https?://[^\s<>\"']+")
    for path in sorted(root.rglob("*.xlsx")):
        try:
            workbook = load_workbook(path, read_only=False, data_only=False, keep_links=True)
        except Exception:
            continue
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows()
            try:
                first_row = next(rows)
            except StopIteration:
                continue

            header_values = [base.clean(cell.value) for cell in first_row]
            normalized_headers = [normalized_header(value) for value in header_values]
            case_id_indexes = [
                index
                for index, header in enumerate(normalized_headers)
                if header in {"id etude de cas", "id cas", "case id"}
            ]

            def inspect_row(row: tuple, row_index: int) -> None:
                values = [base.clean(cell.value) for cell in row]
                source_id = ""
                for index in case_id_indexes:
                    if index < len(values) and CASE_ID_RE.fullmatch(values[index]):
                        source_id = values[index]
                        break
                if not source_id:
                    for index, value in enumerate(values):
                        header = normalized_headers[index] if index < len(normalized_headers) else ""
                        if value and any(token in header for token in (" id ", "numero", "reference")):
                            if CASE_ID_RE.fullmatch(value):
                                source_id = value
                                break

                locator = f"{path.name}:{worksheet.title}:row:{row_index}"
                for cell in row:
                    if cell.hyperlink and cell.hyperlink.target:
                        add(cell.hyperlink.target, locator, source_id)
                    for url in url_rx.findall(base.clean(cell.value)):
                        add(url, locator, source_id)

                # GEIPAN's current case export supplies the stable case ID but
                # not a hyperlink.  Derive only from a validated case ID.
                if source_id:
                    encoded = urllib.parse.quote(source_id, safe="-._~")
                    add(f"{CASE_URL_ROOT}{encoded}", locator, source_id)

            inspect_row(tuple(first_row), 1)
            for row_index, row in enumerate(rows, start=2):
                inspect_row(tuple(row), row_index)

    return sorted(found.values(), key=lambda row: (row.get("source_case_id", ""), row["case_url"]))


def command_plan(args: argparse.Namespace) -> None:
    root = Path(args.live_index_dir)
    urls = discover_with_workbooks(root)
    source_ids = [row.get("source_case_id", "") for row in urls]
    duplicate_source_ids = sorted({source_id for source_id in source_ids if source_id and source_ids.count(source_id) > 1})
    missing_source_ids = sum(1 for source_id in source_ids if not source_id)

    if len(urls) != args.expected_cases or duplicate_source_ids or missing_source_ids:
        diagnostic = {
            "status": "FAIL",
            "expected_case_urls": args.expected_cases,
            "discovered_case_urls": len(urls),
            "missing_source_ids": missing_source_ids,
            "duplicate_source_ids": duplicate_source_ids[:100],
            "first_urls": urls[:20],
            "source_files": [
                str(path.relative_to(root))
                for path in sorted(root.rglob("*"))
                if path.is_file()
            ][:500],
        }
        output = Path(args.output_dir)
        output.mkdir(parents=True, exist_ok=True)
        (output / "URL_DISCOVERY_DIAGNOSTIC.json").write_text(
            json.dumps(diagnostic, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        raise SystemExit(json.dumps(diagnostic, indent=2, ensure_ascii=False))

    output = Path(args.output_dir)
    output.mkdir(parents=True, exist_ok=True)
    (output / "case_urls.json").write_text(
        json.dumps(urls, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    include = []
    for start in range(0, len(urls), args.batch_size):
        include.append(
            {
                "batch_index": start // args.batch_size,
                "start": start,
                "end": min(start + args.batch_size, len(urls)),
                "expected_cases": min(args.batch_size, len(urls) - start),
            }
        )
    matrix = {"include": include}
    (output / "matrix.json").write_text(
        json.dumps(matrix, separators=(",", ":")), encoding="utf-8"
    )
    (output / "PLAN_SUMMARY.json").write_text(
        json.dumps(
            {
                "status": "PASS",
                "case_urls": len(urls),
                "unique_source_case_ids": len(set(source_ids)),
                "batch_size": args.batch_size,
                "batch_count": len(include),
                "source": "validated current-and-legacy GEIPAN source snapshot",
                "url_method": "DERIVED_FROM_CURRENT_CASE_EXPORT_ID_ETUDE_DE_CAS",
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    print(json.dumps(matrix, separators=(",", ":")))


def main() -> None:
    parser = base.parser()
    args = parser.parse_args()
    if args.command == "plan":
        command_plan(args)
    else:
        args.function(args)


if __name__ == "__main__":
    main()
