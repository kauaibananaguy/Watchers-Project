#!/usr/bin/env python3
"""Acquire and census the official GEIPAN downloadable database exports.

This stage preserves the source collection and every physical CSV row. It does
not create a separate public GEIPAN database and does not perform irreversible
canonical matching.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import sqlite3
import time
import urllib.request
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

COLLECTION_ID = "SRC-COLLECTION-GEIPAN"
URLS = {
    "GEIPAN_database_update_page.html": "https://geipan.fr/fr/node/15",
    "GEIPAN_cases.csv": "https://geipan.fr/sites/default/files/Base_de_donn%C3%A9es_des_cas.csv",
    "GEIPAN_testimonies_observations.csv": "https://geipan.fr/sites/default/files/Base_de_donn%C3%A9es_des_t%C3%A9moignages.csv",
    "GEIPAN_table_field_description_2019-01-07.xlsx": "https://geipan.fr/sites/default/files/Description_des_tables_et_champs_de_donn%C3%A9es_de_la_base_du_geipan_2019-01-07.xlsx",
    "GEIPAN_database_history_2019-02-26.pdf": "https://geipan.fr/sites/default/files/2019-02-26_Historique_des_bases_au_GEIPAN.pdf",
    "GEIPAN_statistics_page.html": "https://geipan.fr/fr/stats",
}


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def download(url: str, destination: Path, headers_destination: Path) -> None:
    last: Exception | None = None
    for attempt in range(6):
        try:
            request = urllib.request.Request(
                url,
                headers={
                    "User-Agent": "Mozilla/5.0 Watchers-UFO-Atlas/1.0",
                    "Accept": "*/*",
                },
            )
            with urllib.request.urlopen(request, timeout=300) as response:
                data = response.read()
                if not data:
                    raise RuntimeError(f"Empty response for {url}")
                destination.write_bytes(data)
                headers_destination.write_text(
                    "\n".join(f"{key}: {value}" for key, value in response.headers.items()) + "\n",
                    encoding="utf-8",
                )
                return
        except Exception as exc:  # noqa: BLE001
            last = exc
            if attempt + 1 < 6:
                time.sleep(min(30, 2**attempt))
    raise RuntimeError(f"Unable to download {url}: {last}")


def detect_text(path: Path) -> tuple[str, str]:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace"), "utf-8-replace"


def detect_dialect(text: str) -> csv.Dialect:
    try:
        return csv.Sniffer().sniff(text[:200000], delimiters=";,\t|")
    except csv.Error:
        class Semicolon(csv.excel):
            delimiter = ";"
        return Semicolon()


def clean_text(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text).strip()


def pick(row: dict[str, str], patterns: tuple[str, ...]) -> str:
    for key, value in row.items():
        normalized_key = re.sub(r"[^a-z0-9]+", "_", key.lower()).strip("_")
        if any(re.search(pattern, normalized_key) for pattern in patterns):
            value = clean_text(value)
            if value:
                return value
    return ""


def build(args: argparse.Namespace) -> dict[str, Any]:
    root = Path(args.output_root)
    if root.exists():
        shutil.rmtree(root)
    raw = root / "raw"
    headers = root / "headers"
    output = root / "output"
    raw.mkdir(parents=True)
    headers.mkdir(parents=True)
    output.mkdir(parents=True)

    acquired_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

    inventory: list[dict[str, Any]] = []
    for index, (name, url) in enumerate(URLS.items(), start=1):
        destination = raw / name
        headers_destination = headers / f"{name}.headers.txt"
        print(f"Downloading {name}", flush=True)
        download(url, destination, headers_destination)
        inventory.append(
            {
                "file_id": f"GEIPAN-FILE-{index:03d}",
                "file_name": name,
                "source_url": url,
                "bytes": destination.stat().st_size,
                "sha256": sha256(destination),
                "http_headers": headers_destination.read_text(encoding="utf-8", errors="replace"),
                "acquired_at": acquired_at,
            }
        )

    file_id_by_name = {row["file_name"]: row["file_id"] for row in inventory}
    csv_profiles: dict[str, Any] = {}
    ledger: list[dict[str, Any]] = []
    sequence = 0
    used_ids: dict[str, int] = {}

    for file_name, record_family in (
        ("GEIPAN_cases.csv", "CASE"),
        ("GEIPAN_testimonies_observations.csv", "TESTIMONY_OBSERVATION"),
    ):
        text, encoding = detect_text(raw / file_name)
        dialect = detect_dialect(text)
        reader = csv.DictReader(text.splitlines(), dialect=dialect)
        source_headers = reader.fieldnames or []
        row_count = 0
        blank_rows = 0
        for row_number, row in enumerate(reader, start=2):
            clean = {clean_text(key): clean_text(value) for key, value in row.items() if key is not None}
            if not any(clean.values()):
                blank_rows += 1
                continue
            row_count += 1
            sequence += 1
            native_id = pick(
                clean,
                (
                    r"(^|_)id($|_)",
                    r"numero",
                    r"num_",
                    r"reference",
                    r"cas_id",
                    r"temoignage_id",
                    r"observation_id",
                ),
            )
            base_id = (
                f"GEIPAN-{record_family}-{native_id}"
                if native_id
                else f"GEIPAN-{record_family}-ROW-{row_number:06d}"
            )
            used_ids[base_id] = used_ids.get(base_id, 0) + 1
            source_record_id = (
                base_id
                if used_ids[base_id] == 1
                else f"{base_id}-DUP-{used_ids[base_id]:03d}"
            )
            raw_json = json.dumps(clean, ensure_ascii=False, sort_keys=True)
            ledger.append(
                {
                    "source_collection_id": COLLECTION_ID,
                    "source_sequence": sequence,
                    "source_record_id": source_record_id,
                    "source_native_id": native_id or None,
                    "source_record_family": record_family,
                    "source_file_id": file_id_by_name[file_name],
                    "source_locator": f"{file_name}:row:{row_number}",
                    "original_date_text": pick(clean, (r"date.*observation", r"date.*cas", r"^date$", r"date_debut", r"date_obs")) or None,
                    "original_time_text": pick(clean, (r"heure.*observation", r"^heure$", r"time")) or None,
                    "original_location_text": pick(clean, (r"commune", r"lieu", r"localite", r"departement", r"location")) or None,
                    "title_raw": pick(clean, (r"titre", r"nom.*cas", r"libelle")) or None,
                    "description_raw": pick(clean, (r"recit", r"resume", r"description", r"synthese", r"texte")) or None,
                    "reference_raw": None,
                    "source_attributes_raw": raw_json,
                    "raw_payload": raw_json,
                    "transformation_status": "ACQUIRED_UNTRANSFORMED",
                    "notes": None,
                    "row_sha256": hashlib.sha256(raw_json.encode("utf-8")).hexdigest(),
                }
            )
        csv_profiles[file_name] = {
            "encoding": encoding,
            "delimiter": dialect.delimiter,
            "quotechar": dialect.quotechar,
            "headers": source_headers,
            "header_count": len(source_headers),
            "nonblank_rows": row_count,
            "blank_rows_skipped": blank_rows,
        }

    workbook_profile: dict[str, Any] = {"sheets": []}
    workbook = load_workbook(
        raw / "GEIPAN_table_field_description_2019-01-07.xlsx",
        read_only=True,
        data_only=False,
    )
    for worksheet in workbook.worksheets:
        preview: list[list[str | None]] = []
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if row_index > 12:
                break
            preview.append([None if value is None else str(value) for value in row[:30]])
        workbook_profile["sheets"].append(
            {
                "sheet_name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "preview_rows": preview,
            }
        )

    source_collection = {
        "source_collection_id": COLLECTION_ID,
        "collection_title": "GEIPAN published cases, testimonies, and observations",
        "attribution_name": "GEIPAN — Groupe d’études et d’informations sur les phénomènes aérospatiaux non identifiés",
        "owner_custodian": "Centre national d’études spatiales (CNES)",
        "source_type": "OFFICIAL_GOVERNMENT_DATABASE_EXPORT",
        "source_page_url": "https://geipan.fr/fr/node/15",
        "case_search_url": "https://geipan.fr/fr/recherche/cas",
        "statistics_url": "https://geipan.fr/fr/stats",
        "acquisition_method": "Direct public download from official GEIPAN website",
        "acquired_at": acquired_at,
        "rights_and_attribution_notes": (
            "Official public downloads supplied for database and spreadsheet use. "
            "Witness-confidential fields are withheld by GEIPAN; exported coordinates and ages are generalized for anonymization. "
            "Original French wording and source identifiers must remain preserved."
        ),
        "privacy_notes": (
            "GEIPAN states that witness-confidential process fields are not exported; latitude/longitude are rounded to 0.1 degree, "
            "occupational categories are generalized, and witness ages are binned by decade."
        ),
        "source_file_count": len(inventory),
        "source_record_count": len(ledger),
        "case_row_count": csv_profiles["GEIPAN_cases.csv"]["nonblank_rows"],
        "testimony_observation_row_count": csv_profiles["GEIPAN_testimonies_observations.csv"]["nonblank_rows"],
        "declared_integration_standard": "UFO-ATLAS-INT-STD-1.0.0",
        "declared_gmr_version": "UFO Atlas GMR v1.0.0",
        "transformation_status": "SOURCE_ACQUIRED_AND_CENSUSED",
    }

    ledger_columns = list(ledger[0]) if ledger else []
    with (output / "SOURCE_RECORD_LEDGER.jsonl").open("w", encoding="utf-8") as handle:
        for row in ledger:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")
    with (output / "SOURCE_RECORD_LEDGER.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=ledger_columns)
        writer.writeheader()
        writer.writerows(ledger)

    database = output / "GEIPAN_SOURCE_ACQUISITION_v0.1.0.sqlite"
    connection = sqlite3.connect(database)
    connection.execute("PRAGMA foreign_keys=ON")
    connection.executescript(
        """
        CREATE TABLE source_collection(
          source_collection_id TEXT PRIMARY KEY,
          payload_json TEXT NOT NULL
        );
        CREATE TABLE source_file(
          file_id TEXT PRIMARY KEY,
          file_name TEXT NOT NULL UNIQUE,
          source_url TEXT NOT NULL,
          bytes INTEGER NOT NULL,
          sha256 TEXT NOT NULL,
          acquired_at TEXT NOT NULL,
          http_headers TEXT
        );
        CREATE TABLE source_record_ledger(
          source_collection_id TEXT NOT NULL REFERENCES source_collection(source_collection_id),
          source_sequence INTEGER NOT NULL UNIQUE,
          source_record_id TEXT PRIMARY KEY,
          source_native_id TEXT,
          source_record_family TEXT NOT NULL,
          source_file_id TEXT NOT NULL REFERENCES source_file(file_id),
          source_locator TEXT NOT NULL,
          original_date_text TEXT,
          original_time_text TEXT,
          original_location_text TEXT,
          title_raw TEXT,
          description_raw TEXT,
          reference_raw TEXT,
          source_attributes_raw TEXT,
          raw_payload TEXT NOT NULL,
          transformation_status TEXT NOT NULL,
          notes TEXT,
          row_sha256 TEXT NOT NULL
        );
        CREATE INDEX idx_geipan_source_family ON source_record_ledger(source_record_family);
        CREATE INDEX idx_geipan_native_id ON source_record_ledger(source_native_id);
        CREATE INDEX idx_geipan_date ON source_record_ledger(original_date_text);
        CREATE INDEX idx_geipan_location ON source_record_ledger(original_location_text);
        """
    )
    connection.execute(
        "INSERT INTO source_collection VALUES (?,?)",
        (COLLECTION_ID, json.dumps(source_collection, ensure_ascii=False, sort_keys=True)),
    )
    connection.executemany(
        "INSERT INTO source_file VALUES (?,?,?,?,?,?,?)",
        [
            (
                item["file_id"], item["file_name"], item["source_url"], item["bytes"],
                item["sha256"], item["acquired_at"], item["http_headers"],
            )
            for item in inventory
        ],
    )
    placeholders = ",".join("?" for _ in ledger_columns)
    connection.executemany(
        f"INSERT INTO source_record_ledger VALUES ({placeholders})",
        [tuple(row[column] for column in ledger_columns) for row in ledger],
    )
    connection.commit()
    quick_check = connection.execute("PRAGMA quick_check").fetchone()[0]
    foreign_keys = len(connection.execute("PRAGMA foreign_key_check").fetchall())
    sequence_min, sequence_max, sequence_count = connection.execute(
        "SELECT MIN(source_sequence), MAX(source_sequence), COUNT(DISTINCT source_sequence) FROM source_record_ledger"
    ).fetchone()
    duplicate_ids = connection.execute(
        "SELECT COUNT(*) FROM (SELECT source_record_id FROM source_record_ledger GROUP BY source_record_id HAVING COUNT(*)>1)"
    ).fetchone()[0]
    connection.close()

    validation = {
        "overall_status": "PASS" if (
            quick_check == "ok"
            and foreign_keys == 0
            and duplicate_ids == 0
            and sequence_min == 1
            and sequence_max == len(ledger)
            and sequence_count == len(ledger)
        ) else "FAIL",
        "stage": "SOURCE_ACQUISITION_AND_CENSUS",
        "sqlite_quick_check": quick_check,
        "foreign_key_violations": foreign_keys,
        "duplicate_source_record_ids": duplicate_ids,
        "source_sequence_min": sequence_min,
        "source_sequence_max": sequence_max,
        "source_sequence_distinct_count": sequence_count,
        "source_record_count": len(ledger),
        "case_row_count": csv_profiles["GEIPAN_cases.csv"]["nonblank_rows"],
        "testimony_observation_row_count": csv_profiles["GEIPAN_testimonies_observations.csv"]["nonblank_rows"],
        "source_files": len(inventory),
        "candidate_records_built": 0,
        "gmr_transformation_started": False,
        "next_stage": "SOURCE_FIELD_MAPPING_AND_CANONICAL_CANDIDATE_CONSTRUCTION",
        "known_limitations": [
            "This is an acquisition/census checkpoint, not a completed GMR import module.",
            "The official CSV export intentionally omits confidential witness and internal workflow fields.",
            "Full free-form narratives and some weather/aviation supplemental tables may not be present in the downloadable CSV export.",
        ],
    }
    if validation["overall_status"] != "PASS":
        raise RuntimeError(json.dumps(validation, indent=2, ensure_ascii=False))

    (output / "SOURCE_COLLECTION.json").write_text(json.dumps(source_collection, indent=2, ensure_ascii=False), encoding="utf-8")
    (output / "SOURCE_FILE_INVENTORY.json").write_text(json.dumps(inventory, indent=2, ensure_ascii=False), encoding="utf-8")
    (output / "SOURCE_CENSUS.json").write_text(
        json.dumps(
            {
                "acquired_at": acquired_at,
                "source_collection_id": COLLECTION_ID,
                "source_record_count": len(ledger),
                "case_row_count": csv_profiles["GEIPAN_cases.csv"]["nonblank_rows"],
                "testimony_observation_row_count": csv_profiles["GEIPAN_testimonies_observations.csv"]["nonblank_rows"],
                "source_sequence_start": 1,
                "source_sequence_end": len(ledger),
                "csv_profiles": csv_profiles,
                "geipan_schema_workbook_profile": workbook_profile,
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    (output / "SOURCE_FIELD_HEADERS.json").write_text(json.dumps(csv_profiles, indent=2, ensure_ascii=False), encoding="utf-8")
    (output / "GEIPAN_SCHEMA_WORKBOOK_PROFILE.json").write_text(json.dumps(workbook_profile, indent=2, ensure_ascii=False), encoding="utf-8")
    (output / "VALIDATION_REPORT.json").write_text(json.dumps(validation, indent=2, ensure_ascii=False), encoding="utf-8")

    readme = f"""# GEIPAN official source acquisition — v0.1.0

Status: **{validation['overall_status']}**  
Source collection: `{COLLECTION_ID}`  
Acquired: `{acquired_at}`

This checkpoint preserves the official GEIPAN case CSV, testimony/observation CSV, field-description workbook, database-history PDF, source page, and statistics page. It creates one uninterrupted immutable source ledger covering every nonblank row in both official CSV exports.

## Actual source volume

- Case rows: **{source_collection['case_row_count']:,}**
- Testimony/observation rows: **{source_collection['testimony_observation_row_count']:,}**
- Total physical source rows: **{source_collection['source_record_count']:,}**
- Source files: **{source_collection['source_file_count']}**

## Architectural boundary

This is a source-acquisition checkpoint for the one source-neutral UFO Atlas. It is not a separate public GEIPAN database. No canonical candidates or irreversible matches have yet been created.

## Next stage

Map every GEIPAN source field to the controlling UFO Atlas GMR, preserve French source wording, construct source-neutral case/testimony/observation candidates, write English editorial translations separately, create typed relationships, and prepare duplicate proposals against the latest verified master.
"""
    (output / "README_FIRST.md").write_text(readme, encoding="utf-8")

    # Package the raw source snapshot and its machine-readable ledger.
    release_root = root / "release"
    release_root.mkdir()
    shutil.copytree(raw, release_root / "raw")
    shutil.copytree(output, release_root / "output")
    (release_root / "FRESH_VERIFICATION.json").write_text(
        json.dumps(
            {
                "status": "PASS",
                "sqlite_quick_check": quick_check,
                "foreign_key_violations": foreign_keys,
                "source_record_count": len(ledger),
            },
            indent=2,
        ) + "\n",
        encoding="utf-8",
    )

    package_name = "GEIPAN_OFFICIAL_SOURCE_SNAPSHOT_v0.1.0.zip"
    package = root / package_name
    with zipfile.ZipFile(package, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in sorted(release_root.rglob("*")):
            if path.is_file():
                archive.write(path, path.relative_to(release_root))
    with zipfile.ZipFile(package) as archive:
        bad = archive.testzip()
        if bad:
            raise RuntimeError(f"ZIP integrity failure at {bad}")
    (root / f"{package_name}.sha256").write_text(f"{sha256(package)}  {package_name}\n", encoding="utf-8")

    manifest_files = []
    for path in sorted(release_root.rglob("*")):
        if path.is_file():
            manifest_files.append(
                {
                    "path": str(path.relative_to(release_root)),
                    "bytes": path.stat().st_size,
                    "sha256": sha256(path),
                }
            )
    manifest = {
        "package_id": "UFO-ATLAS-GEIPAN-SOURCE-SNAPSHOT-0.1.0",
        "source_collection_id": COLLECTION_ID,
        "snapshot_date": args.snapshot_date,
        "created_at": acquired_at,
        "package_file": package_name,
        "package_sha256": sha256(package),
        "files": manifest_files,
    }
    (output / "manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")

    print(json.dumps(validation, indent=2, ensure_ascii=False))
    return validation


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-root", default="snapshot")
    parser.add_argument("--snapshot-date", default="2026-09-03")
    args = parser.parse_args()
    build(args)


if __name__ == "__main__":
    main()
